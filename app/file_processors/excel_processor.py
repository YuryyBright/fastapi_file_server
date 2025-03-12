import openpyxl
from .base_processor import FileProcessor

class ExcelProcessor(FileProcessor):
    """
    Processor for Excel files to perform text searches.
    """
    def __init__(self, file_path: str):
        """
        Initialize the Excel processor with the file path.
        :param file_path: Path to the Excel file.
        """
        self.file_path = file_path

    def search(self, keyword: str) -> list[str]:
        """
        Search for a keyword in the Excel file text.
        :param keyword: The keyword to search for.
        :return: A list of lines containing the keyword.
        """
        workbook = openpyxl.load_workbook(self.file_path)
        results = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if keyword in str(cell.value):
                        results.append(str(cell.value))
        return results

    def read(self) -> str:
        """
        Reads the entire content of the Excel file as a string.
        :return: A string representing the content of the Excel file.
        """
        workbook = openpyxl.load_workbook(self.file_path)
        results = ""

        for sheet in workbook.worksheets:
            results += f"Sheet: {sheet.title}\n"
            for row in sheet.iter_rows():
                row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
                results += "\t".join(row_values) + "\n"

        return results
